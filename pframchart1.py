import streamlit as st
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from io import BytesIO, StringIO
import base64

st.markdown("""
<style> 
    #MainMenu{
        visibility:hidden;
    }
    Footer{
        visibility:hidden;
    }
</style>           
""", unsafe_allow_html=True)

st.title("Pfram Chart")

st.markdown(f"[Download the sample file here](https://drive.google.com/uc?id=1Yu7gZh3IWAgM8j7A3GtbMZTJ0NFKuQXK)")



st.write('#')

# Ruta del archivo Excel
archivo_excel = st.file_uploader("Choose a PFRAM file with projects:", type=["xlsm"])

if archivo_excel is not None:
    
    st.write('#')
    mensaje_temporal = st.empty()
    mensaje_temporal.text("Reading and processing data... This may take a moment.")

    try:
        # Cargar el libro de trabajo (workbook)
        libro_trabajo = load_workbook(archivo_excel,  read_only=True, data_only=True)
       
        hojas_disponibles = []
        for nombre_hoja in libro_trabajo.sheetnames:
            if nombre_hoja[0] == 'P' and nombre_hoja[1:].isdigit():
                hoja_actual = libro_trabajo[nombre_hoja]
                valor_fila_31_col_2 = hoja_actual.cell(row=31, column=2).value
                hojas_disponibles.append(valor_fila_31_col_2)

        mensaje_temporal.text("Data processed successfully!")
        st.write('#')

        proyecto_seleccionado = st.multiselect("Choose a project:", hojas_disponibles, placeholder="--Available projects--")

        st.write('#')
        
        if proyecto_seleccionado:
            myDic = {
                241: "Inflows",
                242: "Outflows-Investment",
                243: "Outflows-Operational",
                244: "Outflows-Taxes",
                245: "Outflows-Debt service",
                246: "Outflows-Dividends",
                247: "Equity"
                # Agrega más claves y valores según sea necesario
            }

            datos_seleccionados = st.selectbox("Select your data:", list(myDic.values()), index=None, placeholder="--Choose a parameter--")
            
            fig, ax = plt.subplots()
            plt.style.use("dark_background")

            for proyecto_nombre in proyecto_seleccionado:
                hoja_correspondiente = None
                for nombre_hoja in libro_trabajo.sheetnames:
                    hoja_actual = libro_trabajo[nombre_hoja]
                    valor_fila_31_col_2 = hoja_actual.cell(row=31, column=2).value
                    if valor_fila_31_col_2 == proyecto_nombre:
                        hoja_correspondiente = hoja_actual
                        break

                if hoja_correspondiente:
                    fila_seleccionada = next(key for key, value in myDic.items() if value == datos_seleccionados)
                    valores = [
                        hoja_correspondiente.cell(row=fila_seleccionada, column=col).value
                        for col in range(3, 52)
                        if isinstance(hoja_correspondiente.cell(row=fila_seleccionada, column=col).value, (int, float))
                    ]

                    years = [
                        hoja_correspondiente.cell(row=240, column=col).value
                        for col in range(3, 52)
                        if isinstance(hoja_correspondiente.cell(row=240, column=col).value, (int, float))
                    ] 

                    indice = hoja_correspondiente.cell(row=fila_seleccionada, column=1).value

                    ax.stackplot(years, valores, labels=[f"{proyecto_nombre} - {indice}"], alpha=0.6)

            ax.set_xlabel("Year")
            ax.set_ylabel("Value")
            ax.tick_params(axis='y', labelsize=10)
            ax.set_title(f"Read or compare projects - {datos_seleccionados}")
            ax.legend(fontsize='small', loc='upper center', bbox_to_anchor=(0.5, -0.15), frameon=False)

            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['bottom'].set_visible(True)
            ax.spines['left'].set_visible(True)

            st.pyplot(fig)                   

            buffer = BytesIO()
            plt.savefig(buffer, format='pdf')
            buffer.seek(0)

            b64_file = base64.b64encode(buffer.read()).decode()
            
            href = f'<a href="data:application/octet-stream;base64,{b64_file}" download="stackplot.pdf">Download as PDF</a>'
            st.markdown(href, unsafe_allow_html=True)

            mensaje_temporal = st.empty()
          
         
        else:
            st.warning("Please, select at least one project.")
            

    except Exception as e:
        st.write(f"{e}")
