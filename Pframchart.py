import streamlit as st
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
from io import BytesIO
import base64


st.markdown("""
<style> 
    #MainMenu{
        visibility:hidden;
    }
    Footer{
        visibility:hidden;
    }
</style>           
""", unsafe_allow_html=True)

st.title("Pfram Chart")

# Ruta del archivo Excel
archivo_excel = st.file_uploader("Choose a PFRAM file with projects:", type=["xlsm"])


if archivo_excel is not None:
    
    #st.write("Selected file:", archivo_excel.name)
    st.write('#')
    mensaje_temporal = st.empty()
    mensaje_temporal.text("Reading and processing data... This may take a moment.")

    try:
        # Cargar el libro de trabajo (workbook)
        libro_trabajo = load_workbook(archivo_excel,  read_only=True, data_only=True)
        
        # Crear un nuevo libro de trabajo para los datos
        nuevo_libro = Workbook()
        nuevo_libro.remove(nuevo_libro.active)  # Eliminar la hoja por defecto

        # Obtener hojas que cumplen con el criterio 'P' + número y copiar los datos al nuevo libro
        for nombre_hoja in libro_trabajo.sheetnames:
            if nombre_hoja[0] == 'P' and nombre_hoja[1:].isdigit():
                hoja_actual = libro_trabajo[nombre_hoja]
                nueva_hoja = nuevo_libro.create_sheet(title=nombre_hoja)

                for row in hoja_actual.iter_rows(min_row=1, max_row=hoja_actual.max_row, min_col=1, max_col=hoja_actual.max_column):
                    nueva_hoja.append([cell.value for cell in row])

        # Guardar el nuevo libro de trabajo como myData.xlsx
        nuevo_libro.save("myData.xlsx")

        # Cargar el nuevo libro de trabajo para los datos
        libro_datos = load_workbook("myData.xlsx", read_only=True, data_only=True)
       
        hojas_disponibles = []
        for nombre_hoja in libro_trabajo.sheetnames:
            if nombre_hoja[0] == 'P' and nombre_hoja[1:].isdigit():
                hoja_actual = libro_trabajo[nombre_hoja]
                valor_fila_31_col_2 = hoja_actual.cell(row=31, column=2).value
                hojas_disponibles.append(valor_fila_31_col_2)

        # Actualizar el mensaje después de la tarea
        mensaje_temporal.text("Data processed successfully!")
        
        st.write('#')

        proyecto_seleccionado = st.multiselect("Choose a project:", hojas_disponibles, placeholder="--Available projects--")

        myDic = {
            241: "Inflows",
            242: "Outflows-Investment",
            243: "Outflows-Operational",
            244: "Outflows-Taxes",
            245: "Outflows-Debt service",
            246: "Outflows-Dividends",
            247: "Equity"
            # Agrega más claves y valores según sea necesario
        }

       

         # Selectbox para seleccionar los datos a demostrar
        #datos_seleccionados = st.selectbox("Select your data:", list(myDic.values()), index=0, )
        datos_seleccionados = st.selectbox("Select your data:", list(myDic.values()), index=None,  placeholder="--Choose a parameter--")
       
        if proyecto_seleccionado:
            
            # Crear gráfico de área (stackplot)
            fig, ax = plt.subplots()
            plt.style.use("dark_background")

            for proyecto_nombre in proyecto_seleccionado:
                # Encontrar la hoja correspondiente al proyecto seleccionado
                hoja_correspondiente = None
                for nombre_hoja in libro_trabajo.sheetnames:
                    hoja_actual = libro_trabajo[nombre_hoja]
                    valor_fila_31_col_2 = hoja_actual.cell(row=31, column=2).value
                    if valor_fila_31_col_2 == proyecto_nombre:
                        hoja_correspondiente = hoja_actual
                        break

                if hoja_correspondiente:
                    # Obtener datos de la fila seleccionada, columnas 3 hasta 52
                    fila_seleccionada = next(key for key, value in myDic.items() if value == datos_seleccionados)
                    valores = [
                        hoja_correspondiente.cell(row=fila_seleccionada, column=col).value
                        for col in range(3, 52)
                        if isinstance(hoja_correspondiente.cell(row=fila_seleccionada, column=col).value, (int, float))
                    ]

                    # Obtener años de la fila 240, columnas 3 hasta 52
                    años = [
                        hoja_correspondiente.cell(row=240, column=col).value
                        for col in range(3, 52)
                        if isinstance(hoja_correspondiente.cell(row=240, column=col).value, (int, float))
                    ] 

                    # Obtener índice de la fila seleccionada
                    indice = hoja_correspondiente.cell(row=fila_seleccionada, column=1).value

                    # Agregar los valores al gráfico con el nombre del proyecto
                    ax.stackplot(años, valores, labels=[f"{proyecto_nombre} - {indice}"], alpha=0.6)  # Ajusta alpha según sea necesario

            ax.set_xlabel("Year")
            ax.set_ylabel("Value")
            ax.set_title(f"Compare projects - {datos_seleccionados}")
            ax.legend(fontsize='small', loc='upper center', bbox_to_anchor=(0.5, -0.15), frameon=False)

            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['bottom'].set_visible(True)
            ax.spines['left'].set_visible(True)

            # Mostrar gráfico en Streamlit
            st.pyplot(fig)                   

            # Guardar el gráfico en un BytesIO buffer
            buffer = BytesIO()
            plt.savefig(buffer, format='pdf')
            buffer.seek(0)

            # Codificar el archivo en base64
            b64_file = base64.b64encode(buffer.read()).decode()
            
            # Crear el enlace para descargar el PDF
            href = f'<a href="data:application/octet-stream;base64,{b64_file}" download="stackplot.pdf">Download as PDF</a>'
            st.markdown(href, unsafe_allow_html=True)

            mensaje_temporal = st.empty()
          
         
        else:
            st.warning("Please, select at least one project.")
            

    except Exception as e:
        st.write(f"{e}")

  
